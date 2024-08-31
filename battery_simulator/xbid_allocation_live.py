import functools

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
import datetime as dt
import time
import pprint
import math
import openpyxl
import requests
import urllib3
from openpyxl import Workbook, load_workbook
import logging

from xbid_data_getter import xbid_data

from battery_simulator import Position_xbid
from allocation_optimizer import *
import pickle


class Environment:

    def __init__(self, start_time: dt.datetime, end_time: dt.datetime, soc):
        self.start_time = start_time
        self.end_time = end_time

        # session specific
        self.current_timestamp = None
        self.current_time_from = start_time
        self.current_time_to = end_time

        self.soc = soc


class action_excel_saver:

    def __init__(self, env: Environment):
        self.start_time_string = env.start_time.strftime("%Y%m%d-%H-%M")
        self.end_time_string = env.end_time.strftime("%Y%m%d-%H-%M")
        self.current_row_number, self.initial_allocation_length = 0, None
        self.all_keys = []

        self.workbook_name = f'{os.getcwd()}/production/xbid_excel_logger/{self.start_time_string}_{self.end_time_string}.xlsx'
        if os.path.exists(self.workbook_name):
            self.workbook = load_workbook(self.workbook_name)
            self.worksheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active

    def initialize_interface(self, action):
        keys = np.fromiter(action.keys(), dtype=int)
        self.all_keys = keys
        self.initial_allocation_length = len(keys)
        for index, value in enumerate(keys):
            self.worksheet.cell(row=1, column=2 + index, value=value)
        self.workbook.save(self.workbook_name)
        self.current_row_number = 2

    def add_action(self, env: Environment, action):
        self.worksheet.cell(row=self.current_row_number, column=1, value=dt.datetime.utcnow())
        prices = list(action.values())
        k = 0
        for i, value, in enumerate(self.all_keys):
            if self.all_keys[i] not in list(action.keys()):
                continue
            else:
                self.worksheet.cell(row=self.current_row_number, column=2 + i, value=prices[k])
                k = k + 1
        self.current_row_number = self.current_row_number + 1
        self.workbook.save(self.workbook_name)


class xbid_strategy:

    def __init__(self, start_time, end_time, initial_soc=4, ending_soc=4, quantity=3, action_pair_cost=40, logging_ON=True, called_by_automator=False, *args, **kwargs):
        """
        On going end_time adjustment: What we care for during an end time adjustment (the backbone variables of the algorithm):
        the true old ask and true old bid dictionaries. The current allocation, the position object, the soc state (all these in correspondance to the new ask and new dicts)
        ALLOCATION should be converted to dict, keys being the market that correpsond to the buyorsell
        (now its a list), in the style of the position object. Will be more transparent, for what is going on

        """
        self.action_pair_cost_list = None
        self.extend_end_time = False
        self.interrupt_running = False
        self.logging_ON = logging_ON
        self.called_by_automator = called_by_automator
        self.data = xbid_data()
        self.position = Position_xbid()
        self.start_time = start_time
        self.end_time = end_time
        self.cut_threshold = 7  # variable to check up to how many minutes before market closure to trade
        self.first_allocation = None
        self.bid, self.ask = None, None
        self.sleep_time = 60

        self.max_capacity = 10
        self.initial_soc = initial_soc
        self.ending_soc = ending_soc
        self.action_cost = 1
        self.current_soc = initial_soc
        self.soc_history = []
        self.allocation_quantity = quantity
        self.action_pair_cost = action_pair_cost  # chosen for 3 mw
        self.action_pair_cost_list = {'low': 15, 'high': 35}

        # allocation_profit part: current total settling amount given allocation
        # trading_profit part: profit/loss from closure of positions, when updating
        self.allocation = []
        self.allocation_profit, self.trading_profit, self.current_settling_profit, self.total_current_profit = 0, [0], [], 0
        self.allocation_cost = 0
        self.true_old_bid, self.true_old_ask = {}, {}
        self.allocation_history, self.actions_history = [], {}
        self.total_settled_profit = None
        self.total_projected_profit = None

        # variables for checking various statuses
        self.iteration_profits = []  # save iteration profits, to measure the std, and try to capture potentially faulty bid ask requests/ or bid ask handling
        self.potentially_faulty_data = []
        self.live_time = None
        self.current_contracts = None
        self.object_file = None

    def __repr__(self):
        return f'{self.start_time.strftime("%Y-%m-%d")}: {self.start_time.hour} - {self.end_time.hour} | init_state: {self.initial_soc}'

    def custom_allocation_initializer(self):
        # self.allocation, self.position, self.true_old_ask, self.true_old_bid
        pass

    def interrupt_running(self):
        self.interrupt_running = True

    def run(self):
        env = Environment(self.start_time, self.end_time, self.current_soc)
        excel_saver = action_excel_saver(env)
        pickle_dir = f'{os.getcwd()}\production\\xbid_pickled_objects'
        self.object_file = f'{pickle_dir}\\{env.start_time.strftime("%Y-%m-%d %H-%M")}_{env.end_time.strftime("%Y-%m-%d %H-%M")}.txt'
        env.current_timestamp = self.get_current_timestamp()
        env.current_time_from, env.current_time_to = self.update_current_period(env)
        if self.first_allocation is not None:
            self.extend_end_time = True

        while env.current_timestamp < self.end_time - dt.timedelta(minutes=30):

            bid, ask, self.current_contracts = self.get_bid_ask(env)
            self.bid, self.ask = bid.copy(), ask.copy()

            # if a key settles, we remove the respective position from position, and get a new charge level.
            env.current_timestamp = self.get_current_timestamp()
            env.current_time_from, env.current_time_to = self.update_current_period(env)
            if len(ask) > len(self.true_old_ask) and self.extend_end_time is True:
                self.updates_dicts_allocation_and_position(ask, bid)
                self.extend_end_time = False
            if self.position.check_and_settle_position(bid, self.allocation):
                env.soc = self.get_current_soc()

            if self.interrupt_running is True:
                return None
            if bid != {} and ask != {} and len(bid) > 1:
                if self.first_allocation is None:
                    initial_allocation, clean_profit, number_of_actions = get_initial_allocation(ask, bid, env.soc, self.ending_soc, 1, action_pair_cost=self.action_pair_cost)
                    # assume cost per cycle is 1 eur
                    if sum(np.abs(initial_allocation)) != 0 and clean_profit >= number_of_actions:
                        # total allocation action cost
                        allocation_cost = np.count_nonzero(list(initial_allocation)) * self.action_pair_cost / 2
                        print(f'{self.start_time.strftime("%Y-%m-%d")}: {self.start_time.hour} - {self.end_time.hour} | init_state: {self.initial_soc}')
                        print('allocation Initizialized, with profit: ', clean_profit, ' + actions cost: ', allocation_cost, '= ', clean_profit + allocation_cost, ',  number of actions: ',
                              number_of_actions)
                        self.first_allocation = True
                        self.allocation = initial_allocation
                        self.allocation_history.append(self.allocation)
                        self.allocation_profit = clean_profit + allocation_cost
                        self.allocation_cost = allocation_cost
                        self.true_old_bid, self.true_old_ask = bid, ask
                        self.position.add_change_position(bid, ask, self.allocation)
                        action = dict(zip(ask.keys(), initial_allocation))
                        self.actions_history[env.current_timestamp] = action
                        if self.logging_ON:
                            excel_saver.initialize_interface(action)
                            excel_saver.add_action(env, action)

                else:
                    [[new_allocation, clean_combined_profit, number_of_actions]], [allocation_profit, trading_profit], [next_old_ask, next_old_bid, trimmed_old_allocation] = \
                        get_new_allocation(ask, bid, self.true_old_ask, self.true_old_bid, self.allocation, env.soc, self.ending_soc, action_pair_cost=self.action_pair_cost)
                    allocation_cost = np.count_nonzero(new_allocation) * self.action_pair_cost / 2
                    combined_profit_with_cost = allocation_cost + clean_combined_profit  # this has to be equal with: allocation_profit + trading_profit that dont incorporate the actions

                    #  in the updation the true cost of the number of actions, is the difference of the amounts of actions of the new_allocation and the old_allocation
                    #  Each time, the first time  we update the allocation, where a new settle amount has been introduced, we need to add the difference of the settling amounts.
                    #  each allocation length will correspond to different settled_allocation_profits. Basically with the settling difference
                    #  we are comparing the different allocation profits correctly.
                    #  In each allocation part profit we have to add the respective current settling_amount. This materializes in the updation inequality check
                    try:
                        settled_profit = sum(self.position.settled.values())
                        settled_cost = np.count_nonzero(list(self.position.settled_type.values())) * self.action_pair_cost / 2
                    except:
                        settled_profit = 0
                        settled_cost = 0
                    settling_difference = 0
                    if len(new_allocation) < len(self.allocation):
                        # diff_length: to find the correct settled amount difference for the new potential and old allocation
                        difference_length = len(self.allocation) - len(new_allocation)
                        settling_difference_actions_cost = np.count_nonzero(list(self.position.settled_type.values())[-difference_length:]) * self.action_pair_cost / 2
                        settling_difference = sum(self.position.settled.values()) - sum(list(self.position.settled.values())[:-difference_length])
                        settling_difference = settling_difference + settling_difference_actions_cost

                    self.print_interval_results(allocation_profit, trading_profit, settled_profit, allocation_cost, settled_cost, number_of_actions)

                    # trimmed_old_allocation is self.allocation before new_update, without the expired keys. so we can compare same list length, that corresponds to the same keys.
                    # (basically allocation_part + trading_part + settled_profit) of the new allocation > (basically allocation_part + trading_part + settled_profit) of old allocation
                    # so if you do the math we see settling_difference = settled_profit of the new allocation  - settled_profit of old_allocation. If you insert it above
                    # is what we get in the below if statement
                    # percentage check: e.g. if we take 5 actions, profit has to be 5% than old_allocation_profit, because in real time we expect error of api vs executed price
                    # fixed check: attribute 1eur for each non-zero value of the action dictionary, arbitrarily
                    action = dict(zip(ask.keys(), list(np.subtract(new_allocation, trimmed_old_allocation))))
                    my_updation_cost = np.count_nonzero(list(action.values()))/2
                    custom_check_value = clean_combined_profit + settling_difference
                    margin_threshold = (1 + 0.0165 * my_updation_cost)
                    if self.allocation_profit - self.allocation_cost < 0:
                        margin_threshold = 1 / margin_threshold
                    if (allocation_profit + settled_profit) > 150:
                        trading_threshold = (np.sqrt(np.abs(allocation_profit + settled_profit)))
                    else:
                        trading_threshold = (np.sqrt(np.abs(allocation_profit + settled_profit))) + 2.5

                    if custom_check_value > (self.allocation_profit - self.allocation_cost) * margin_threshold and list(
                            trimmed_old_allocation) != list(new_allocation):
                        clean_iteration_profit = np.round(allocation_profit - allocation_cost + trading_profit + settled_profit - settled_cost, 2)
                        if trading_profit > trading_threshold or \
                            (custom_check_value > ((self.allocation_profit - self.allocation_cost) * margin_threshold + my_updation_cost) and trading_profit == 0) or \
                            (custom_check_value > ((self.allocation_profit - self.allocation_cost) * margin_threshold + my_updation_cost) and (0.65*trading_threshold < np.abs(trading_profit) < 0.8*trading_threshold)):
                            # this self.allocation will eventually turn to be a trimmed_old_allocation for the next update
                            self.allocation = new_allocation
                            env.soc = self.get_current_soc()
                            self.allocation_profit = allocation_profit
                            self.allocation_cost = allocation_cost
                            self.allocation_history.append([self.allocation, f'soc_now: {env.soc}'])
                            self.trading_profit.append(trading_profit)
                            self.true_old_bid, self.true_old_ask = next_old_bid, next_old_ask
                            if len(bid) == len(self.allocation):
                                self.position.add_change_position(bid, ask, self.allocation)
                            else:
                                print(bid, '\n', self.allocation)
                            self.actions_history[env.current_timestamp] = action
                            if self.logging_ON:
                                excel_saver.add_action(env, action)
                            self.print_interval_results(allocation_profit, trading_profit, settled_profit, allocation_cost, settled_cost, number_of_actions)
                            print(f'{self.start_time.strftime("%Y-%m-%d")}: {self.start_time.hour} - {self.end_time.hour} --> '
                                  f'allocation Updated, number_of_actions = ', number_of_actions, '\n')
                            print('actions_to_take', action, '\n')

            if self.allocation_profit:  # We don't care to update the true_old_ask and true_old_bid values if we dont update the allocation, because we dont care about the prices where allocation is 0
                self.total_projected_profit = self.allocation_profit + sum(self.trading_profit)
                self.total_settled_profit = sum(self.position.settled.values()) + sum(self.trading_profit)
            if self.logging_ON:
                self.save_object_to_file()
            time.sleep(self.sleep_time)
            env.current_timestamp = self.get_current_timestamp()
            self.live_time = dt.datetime.utcnow()

        if self.interrupt_running is False:   # create the results
            self.position.settle_remaining_position(self.allocation)
            self.total_settled_profit = sum(self.position.settled.values()) + sum(self.trading_profit)
            env.current_timestamp = self.get_current_timestamp()
            # should remove the if len(bid) == 1
            settled_allocation, settled_allocation_type, settled_allocation_profit, total_trading_profit, actions_history = self.get_results()
            total_profit = settled_allocation_profit + total_trading_profit
            actions_pairs = np.count_nonzero(list(self.position.settled_type.values())) / 2
            print('settled_allocation_profit + trading_profit = ', settled_allocation_profit, '+', total_trading_profit, "=",
                  total_profit, 'with', actions_pairs, 'action_pairs \n')
            if self.logging_ON:
                self.results_to_logs(total_profit, actions_pairs)
            if self.called_by_automator:
                ret = (total_profit, actions_pairs)
                return ret
        # return [settled_allocation_profit, total_trading_profit, settled_allocation, settled_allocation_type]

    def get_results(self):
        settled_allocation = self.position.settled
        settled_allocation_type = self.position.settled_type
        settled_allocation_profit = sum(self.position.settled.values())
        total_trading_profit = sum(self.trading_profit)
        actions_history = self.actions_history
        return settled_allocation, settled_allocation_type, settled_allocation_profit, total_trading_profit, actions_history

    def get_current_soc(self):
        # the self.settled_type variable of the position object will contained the settled allocation which can be -1, 0 , 1. if 1 we buy, -1 we sell
        soc = self.initial_soc + sum(list(self.position.settled_type.values()))
        if soc > 4 or soc < 0:
            print('soc value is wrong now:', soc)
            print('bid:', self.bid, '\n', 'ask:', self.ask)
        self.soc_history.append(soc)

        return soc

    def get_current_timestamp(self):
        if dt.datetime.utcnow() > self.start_time:
            current_timestamp = dt.datetime.utcnow()
        else:
            current_timestamp = self.start_time

        return current_timestamp

    def get_bid_ask(self, env):
        bid, ask, contracts, data_checks = None, None, None, None
        flag, exception_raised = True, False  # flag is used to check if an exception is raised or faulty data are created somehow
        exceptions_counter = 0
        logging.basicConfig(filename='xbid_data_errors_MAIN.log')
        while flag is True:
            try:
                if exception_raised is True:
                    self.data = xbid_data()
                env.current_timestamp = self.get_current_timestamp()
                env.current_time_from, env.current_time_to = self.update_current_period(env)
                start_time, end_time = env.current_time_from, env.current_time_to
                bid, ask, contracts, = self.data.get_best_bid_ask_profile(start_time, end_time, quantity=self.allocation_quantity)
                # the last 2 checks is to check if we have taken for a key a zero value, which means wrong calculation or faulty request, not sure ye
                data_checks = [len(bid) == len(contracts), [isinstance(x, float) for x in list(bid.values())] == [True] * len(bid),
                               [isinstance(x, float) for x in list(ask.values())] == [True] * len(ask)]
                if data_checks == [True] * len(data_checks):
                    flag = False
            except (Exception, requests.exceptions.RequestException) as ee:
                exception_raised = True
                exceptions_counter += 1
                logging.exception(ee)
                if exceptions_counter == 1:
                    del self.data
                try:
                    del self.data
                except Exception as ee:
                    pass
                time.sleep(10)
                continue

        if exceptions_counter != 0:
            print('\n xbid data exception was raised :', exceptions_counter, 'times')

        return bid, ask, contracts

    def update_current_period(self, env):
        # env is used mainly to check which ask bid keys to load in each iteration.
        # this returns the time interval for the ask bid , in each iteration.
        c, start = 8, None  # minutes to stop before trading
        if env.current_timestamp <= self.start_time - dt.timedelta(minutes=17):
            env.current_timestamp = self.start_time
        if 0 <= env.current_timestamp.minute < 15 - c:
            start = env.current_timestamp.replace(minute=15, second=0)
        elif env.current_timestamp.minute < 30 - c:
            start = env.current_timestamp.replace(minute=30, second=0)
        elif env.current_timestamp.minute < 45 - c:
            start = env.current_timestamp.replace(minute=45, second=0)
        elif env.current_timestamp.minute < 60 - c:
            start = (env.current_timestamp + dt.timedelta(hours=1)).replace(minute=0, second=0)
        elif env.current_timestamp.minute < 60:
            start = (env.current_timestamp + dt.timedelta(hours=1)).replace(minute=15, second=0)

        end = self.end_time

        return start - dt.timedelta(seconds=1), end

    def save_object_to_file(self):
        if self.first_allocation is not None:
            open(self.object_file, 'wb').close()
        write_file = open(self.object_file, 'wb')
        pickle.dump(self, write_file)
        write_file.close()
        # object_from_file = open(self.object_file, 'rb')

    @staticmethod
    def print_interval_results(allocation_profit, trading_profit, settled_profit, cost, settled_cost, number_of_actions):
        gross_profit = np.round(allocation_profit + trading_profit + settled_profit, 2)
        clean_profit = np.round(allocation_profit - cost + trading_profit + settled_profit - settled_cost, 2)
        return print(f'allocation_part: {allocation_profit} + trading_part: {trading_profit} + settled_part: {np.round(settled_profit, 2)}  =  {gross_profit} '
                     f'- action_cost: {cost} - settled_cost: {settled_cost} = {clean_profit},  number of actions:  {number_of_actions}')

    def results_to_logs(self, total_profit, action_pairs):
        # creation - data to save
        start = pd.Timestamp(self.start_time).tz_localize('UTC').tz_convert('CET')
        hours_difference = (self.end_time - self.start_time).seconds/3600
        if start.hour >= 21 or ((self.end_time.day != self.start_time.day) and hours_difference >= 20):
            start_day = self.start_time+dt.timedelta(days=1)
        else:
            start_day = self.start_time
        date = f'{start_day.day}/{start_day.month}/{start_day.year}'
        if self.initial_soc == 0:
            initial_state = 'discharged'
        else:
            initial_state = 'charged'
        time_interval_utc_00 = f'{self.start_time.hour}:00 - {self.end_time.hour}:00'
        quantity = self.allocation_quantity
        cycles = (action_pairs * quantity)/(self.max_capacity*4)
        total_profit_quantity_adjusted = total_profit * quantity/4

        # excel file handling
        workbook_name = f'{os.getcwd()}\production\\xbid_allocation_results_data.xlsx'
        try:
            workbook = load_workbook(workbook_name)
            worksheet = workbook['specific_xbid_intervals']
            worksheet.append([date, initial_state, time_interval_utc_00, quantity, total_profit_quantity_adjusted, action_pairs, cycles])
            workbook.save(workbook_name)
        except Exception as e:
            print(e)

    def updates_dicts_allocation_and_position(self, ask, bid):
        allocation = list(self.allocation.copy())
        for key in ask:
            if key not in self.true_old_ask:
                self.true_old_ask[key] = ask[key]
                self.true_old_bid[key] = bid[key]
            self.position.position[key] = 0
        for i in range(len(ask) - len(self.allocation)):
            allocation.append(0)
        self.allocation = np.array(allocation)

    def current_action_cost(self):
        allocations_made = len(self.allocation_history)
        action_pair_cost_list = list(self.action_pair_cost_list.values())
        if self.action_pair_cost > np.mean(action_pair_cost_list):
            starting_index = 0
        else:
            starting_index = 1
        if allocations_made % 2 == 1:
            return action_pair_cost_list[0 + starting_index]
        else:
            return action_pair_cost_list[1 - starting_index]

# give datetime in uct FORMAT, so if we want 00:00 to 1 am cet we give 23:00 - 00:00 pm utc time (if day light is on)
# time_from = dt.datetime(2022, 12, 16, 13, 0)
# time_to = dt.datetime(2022, 12, 16, 16, 0)
#
# a = xbid_strategy(time_from, time_to, 4,4)
# a.run()
# x = 2
